# pages/admin.py
import io
import calendar
import datetime
import locale
import requests
from datetime import timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple
import numpy as np
import pandas as pd
import bcrypt
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from utils._config import HEADERS_USERS
from utils._database import get_db

# ====================================================
# 1. CONFIGURAÇÃO DA PÁGINA
# ====================================================
st.set_page_config(
    page_title="Painel Admin | TOTALE",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ====================================================
# 2. PROTEÇÃO DE ACESSO
# ====================================================
if not st.session_state.get("authenticated", False):
    st.switch_page("streamlit_app.py")

info_logado = st.session_state.get("user_info", {}) or {}
nome_logado = str(info_logado.get("tecnico", "")).strip().upper()
login_logado = str(info_logado.get("login", "")).strip().upper()
user_logado = str(info_logado.get("user", "")).strip().upper()
perfil_logado = str(info_logado.get("perfil", "")).strip().upper()

ADMIN_IDENTIFIERS = {
    "ADMIN",
    "ADM",
    "ADMINISTRADOR",
    "GESTOR",
    "MANAGER",
    "SUPERVISOR",
    "COORDENADOR",
}
is_admin = perfil_logado in ADMIN_IDENTIFIERS or login_logado.startswith("ADM")

if not is_admin:
    st.error(
        "🔒 **Acesso Negado** — Esta área é restrita a administradores e gestores."
    )
    st.info(f"Seu perfil atual: **{perfil_logado or 'TÉCNICO'}**.")
    if st.button("⬅️ Voltar para Produção", type="primary"):
        st.switch_page("pages/producao.py")
    st.stop()

# ====================================================
# 3. IMPORTAÇÃO DO DESIGN SYSTEM
# ====================================================
try:
    from components.componentes import (
        aplicar_estilo,
        aplicar_tema_claro,
        render_hero_totale_1,
        render_insight,
        render_section_header,
        injetar_css_menu_nomes,  # noqa: F401
    )
except ImportError:
    st.error("⚠️ Módulo 'componentes.py' não encontrado.")
    st.stop()


def configurar_locale() -> None:
    for loc in ("pt_BR.UTF-8", "Portuguese_Brazil.1252", "pt_BR"):
        try:
            locale.setlocale(locale.LC_TIME, loc)
            return
        except locale.Error:
            continue


configurar_locale()


# ====================================================
# 4. CSS UNIFICADO
# ====================================================
def injetar_css_admin() -> None:
    st.markdown(
        """
        <style>
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0F172A 0%, #1E293B 100%) !important;
            border-right: 1px solid #334155 !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] li:first-child { display: none !important; }
        [data-testid="stSidebarNav"] a[href*="consultivo"] { display: none !important; }
        [data-testid="stSidebarNav"] a[href*="producao"] { display: none !important; }
        [data-testid="stSidebarNav"] a[href*="admin"] span { font-size: 0 !important; }
        [data-testid="stSidebarNav"] a[href*="admin"] span::before {
            content: "🛡️ Admin" !important; font-size: 14px !important;
            font-weight: 700 !important; color: #F8FAFC !important;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a {
            background: rgba(255,255,255,0.05) !important;
            border: 1px solid #334155 !important;
            border-radius: 8px !important;
            margin: 4px 16px !important;
            padding: 10px 14px !important;
            transition: all 0.2s ease;
        }
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[aria-current="page"],
        section[data-testid="stSidebar"] [data-testid="stSidebarNav"] a[aria-selected="true"] {
            background: linear-gradient(90deg, rgba(249,115,22,0.15) 0%, rgba(249,115,22,0.05) 100%) !important;
            border: 1px solid #F97316 !important;
            border-left: 4px solid #F97316 !important;
        }
        section[data-testid="stSidebar"] .admin-profile-card,
        section[data-testid="stSidebar"] .admin-profile-card div,
        section[data-testid="stSidebar"] .admin-profile-card p,
        section[data-testid="stSidebar"] .admin-profile-card span { color: #F8FAFC !important; }
        section[data-testid="stSidebar"] .admin-profile-card .admin-profile-label { color: #F97316 !important; }
        section[data-testid="stSidebar"] .admin-profile-card .admin-profile-login { color: #CBD5E1 !important; }
        section[data-testid="stSidebar"] .admin-profile-card .admin-profile-login-value { color: #FFFFFF !important; }
        section[data-testid="stSidebar"] .admin-profile-card .admin-profile-status { color: #FCA5A5 !important; }
        section[data-testid="stSidebar"] .admin-profile-card .badge-admin { color: #FFFFFF !important; }
        section[data-testid="stSidebar"] .admin-profile-card .admin-profile-avatar { color: #FFFFFF !important; }
        .btn-logout button {
            background: linear-gradient(180deg, #334155 0%, #1E293B 100%) !important;
            border: 1px solid #475569 !important;
            color: #F8FAFC !important;
            font-weight: 600 !important;
            border-radius: 8px !important;
            margin-top: 10px !important;
        }
        .btn-logout button:hover { background: #475569 !important; border-color: #F97316 !important; }
        .card-admin { position: relative; cursor: default; }
        .tooltip-premium {
            visibility: hidden; background-color: #0F172A; color: #F8FAFC;
            text-align: center; border-radius: 8px; padding: 8px 12px;
            position: absolute; z-index: 999; bottom: 115%; left: 50%;
            transform: translateX(-50%); opacity: 0;
            transition: opacity 0.3s ease, bottom 0.3s ease;
            font-size: 11px; font-weight: 500; min-width: 200px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.3); pointer-events: none;
        }
        .tooltip-premium::after {
            content: ""; position: absolute; top: 100%; left: 50%;
            margin-left: -6px; border-width: 6px; border-style: solid;
            border-color: #0F172A transparent transparent transparent;
        }
        .card-admin:hover .tooltip-premium { visibility: visible; opacity: 1; bottom: 105%; }
        [data-testid="stMain"] [data-baseweb="input"],
        [data-testid="stMain"] [data-testid="stDateInput"] > div > div,
        [data-testid="stMain"] div[data-baseweb="select"] > div {
            background-color: #FFFFFF !important;
            border: 1px solid #CBD5E1 !important;
            border-radius: 8px !important;
        }
        [data-testid="stMain"] input {
            color: #0F172A !important; font-weight: 600 !important; font-size: 14px !important;
        }
        .badge-admin {
            display: inline-flex; align-items: center; gap: 6px;
            background: linear-gradient(90deg, #F97316, #EA580C);
            color: white; font-size: 10px; font-weight: 800;
            padding: 3px 10px; border-radius: 20px;
            letter-spacing: 1px; text-transform: uppercase;
        }
        .loading-totale {
            background: #F8FAFC; border: 1px solid #E2E8F0; border-left: 4px solid #F37C04;
            padding: 16px 20px; border-radius: 10px; margin-bottom: 20px;
            display: flex; align-items: center; justify-content: space-between;
        }
        .stTabs [data-baseweb="tab-list"] {
            gap: 8px; background-color: transparent;
            padding-bottom: 2px; border-bottom: 2px solid #E2E8F0;
        }
        .stTabs [data-baseweb="tab"] {
            background-color: #F8FAFC !important; border: 1px solid #E2E8F0 !important;
            border-bottom: none !important; border-radius: 8px 8px 0 0 !important;
            padding: 12px 24px !important; font-weight: 700 !important;
            color: #64748B !important; transition: all 0.3s ease-in-out !important;
        }
        .stTabs [data-baseweb="tab"]:hover {
            background-color: #E2E8F0 !important; color: #0F172A !important;
        }
        .stTabs [aria-selected="true"] {
            background: linear-gradient(90deg, #012869 0%, #1E3A8A 100%) !important;
            color: #FFFFFF !important; border: 1px solid #012869 !important;
            border-bottom: none !important;
            box-shadow: 0 -4px 10px rgba(1, 40, 105, 0.2) !important;
        }
        .stTabs [data-baseweb="tab-highlight"] { display: none !important; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_sidebar_admin() -> None:
    partes = nome_logado.split()
    if len(partes) >= 2:
        iniciais = partes[0][0] + partes[1][0]
    elif len(partes) == 1:
        iniciais = partes[0][:2]
    else:
        iniciais = "AD"

    with st.sidebar:
        st.markdown(
            f"""
            <div class="admin-profile-card" style="background: linear-gradient(145deg, #1E293B 0%, #0F172A 100%);
                        border: 1px solid #334155; border-left: 4px solid #F97316;
                        border-radius: 12px; padding: 16px; margin: 16px;">
                <div class="admin-profile-label" style="font-size: 11px; font-weight: 800; letter-spacing: 1.5px; color: #F97316; margin-bottom: 12px;">
                    ⚡ TOTALE · ADMIN
                </div>
                <div class="admin-profile-avatar" style="width: 52px; height: 52px; border-radius: 50%;
                            background: linear-gradient(135deg, #F97316 0%, #DC2626 100%);
                            color: white; display: flex; align-items: center; justify-content: center;
                            font-size: 18px; font-weight: 800; margin-bottom: 12px;
                            box-shadow: 0 4px 10px rgba(249,115,22,0.3);">
                    {iniciais}
                </div>
                <p style="color: #F8FAFC; font-weight: 800; font-size: 14px; margin: 0 0 4px 0;">
                    {nome_logado or 'ADMINISTRADOR'}
                </p>
                <div style="margin-bottom: 12px;">
                    <span class="badge-admin">🛡️ {perfil_logado or 'ADMIN'}</span>
                </div>
                <div style="display: flex; flex-direction: column; gap: 6px; margin-bottom: 16px;">
                    <div class="admin-profile-login" style="font-size: 12px; color: #94A3B8;">
                        Login: <span class="admin-profile-login-value" style="background: #334155; color: #E2E8F0; padding: 2px 6px; border-radius: 4px; font-family: monospace; font-weight: 600;">{login_logado or 'N/A'}</span>
                    </div>
                </div>
                <div class="admin-profile-status" style="background: rgba(239,68,68,0.15); border: 1px solid #EF4444; color: #FCA5A5;
                            padding: 4px 12px; border-radius: 20px; font-size: 11px; font-weight: 700;
                            display: inline-flex; align-items: center; gap: 6px;">
                    <div style="width: 6px; height: 6px; background: #EF4444; border-radius: 50%;"></div>
                    Sessão Admin
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        st.markdown('<div class="btn-logout">', unsafe_allow_html=True)
        if st.button("🚪 Encerrar Sessão", use_container_width=True):
            st.session_state["authenticated"] = False
            st.switch_page("streamlit_app.py")
        st.markdown("</div>", unsafe_allow_html=True)


# ====================================================
# 5. CARREGAMENTO DAS BASES
# ====================================================
SPREADSHEET_ID_PRODUCAO = "11Dp9WdZYUrT_LBvfo07Mi8muKXZykU7v"
SPREADSHEET_ID_ATIVOS = "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg"
SPREADSHEET_ID_CONSULTIVO = "1YOWJ0HuGcEP2vJaZwl2kcgrtNgsoMBDs"


@st.cache_data(ttl=600, show_spinner=False)
def carregar_base_producao() -> pd.DataFrame:
    urls_xlsx = [
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID_PRODUCAO}/export?format=xlsx"
    ]
    for url in urls_xlsx:
        try:
            resp = requests.get(url, timeout=35)
            if resp.status_code == 200 and len(resp.content) > 100:
                excel_file = pd.read_excel(
                    io.BytesIO(resp.content), sheet_name=None, engine="openpyxl"
                )
                dfs = []
                for sheet_name, df_sheet in excel_file.items():
                    if not df_sheet.empty and df_sheet.shape[1] > 1:
                        cols_upper = [str(c).upper().strip() for c in df_sheet.columns]
                        if not any(
                            k in cols_upper
                            for k in ["ORIGEM", "TIPO", "SISTEMA", "FONTE", "TIPO_OS"]
                        ):
                            df_sheet["ORIGEM"] = sheet_name
                        dfs.append(df_sheet)
                if dfs:
                    return pd.concat(dfs, ignore_index=True)
        except Exception:
            pass

    urls_csv = [
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID_PRODUCAO}/export?format=csv",
        f"https://drive.google.com/uc?export=download&id={SPREADSHEET_ID_PRODUCAO}",
    ]
    for url in urls_csv:
        try:
            resp = requests.get(url, timeout=30)
            if resp.status_code == 200 and len(resp.content) > 100:
                for sep_val, enc in ((None, "utf-8"), (";", "latin-1"), (",", "utf-8")):
                    try:
                        eng = "python" if sep_val is None else "c"
                        df = pd.read_csv(
                            io.BytesIO(resp.content),
                            sep=sep_val,
                            engine=eng,
                            encoding=enc,
                        )
                        if not df.empty and df.shape[1] > 1:
                            return df
                    except Exception:
                        continue
        except Exception:
            continue
    return pd.DataFrame()


@st.cache_data(ttl=600, show_spinner=False)
def carregar_lista_ativos() -> pd.DataFrame:
    urls = [
        f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID_ATIVOS}/export?format=csv",
        f"https://drive.google.com/uc?export=download&id={SPREADSHEET_ID_ATIVOS}",
    ]
    for url in urls:
        try:
            resp = requests.get(url, timeout=30)
        except requests.RequestException:
            continue
        if resp.status_code != 200 or len(resp.content) <= 100:
            continue
        content = resp.content
        for sep_val, enc in ((None, "utf-8"), (";", "latin-1"), (",", "utf-8")):
            try:
                eng = "python" if sep_val is None else "c"
                df = pd.read_csv(
                    io.BytesIO(content), sep=sep_val, engine=eng, encoding=enc
                )
                if not df.empty and df.shape[1] > 1:
                    return df
            except Exception:
                continue
        try:
            df = pd.read_excel(io.BytesIO(content), engine="openpyxl")
            if not df.empty and df.shape[1] > 1:
                return df
        except Exception:
            continue
    return pd.DataFrame()


@st.cache_data(ttl=600, show_spinner=False)
def carregar_base_consultivo() -> pd.DataFrame:
    file_id = SPREADSHEET_ID_CONSULTIVO
    urls = [
        f"https://drive.google.com/uc?export=download&id={file_id}&confirm=t",
        f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx",
        f"https://drive.usercontent.google.com/download?id={file_id}&export=download",
        f"https://drive.google.com/uc?export=download&id={file_id}",
        f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv",
    ]
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/118.0.0.0 Safari/537.36"
            )
        }
    )
    for url in urls:
        try:
            resp = session.get(url, timeout=30, allow_redirects=True)
            if resp.status_code == 200 and len(resp.content) > 100:
                content = resp.content
                if content.strip().startswith(
                    b"<!DOCTYPE html"
                ) or content.strip().startswith(b"<html"):
                    continue
                try:
                    excel_file = pd.read_excel(
                        io.BytesIO(content), sheet_name=None, engine="openpyxl"
                    )
                    dfs = []
                    for sheet_name, df_sheet in excel_file.items():
                        if not df_sheet.empty and df_sheet.shape[1] > 1:
                            df_sheet["_ABA_ORIGEM"] = sheet_name
                            dfs.append(df_sheet)
                    if dfs:
                        return pd.concat(dfs, ignore_index=True)
                except Exception:
                    pass
                for sep_val in (None, ";", ",", "\t"):
                    for enc in ("utf-8", "latin-1", "cp1252"):
                        try:
                            eng = "python" if sep_val is None else "c"
                            df = pd.read_csv(
                                io.BytesIO(content),
                                sep=sep_val,
                                engine=eng,
                                encoding=enc,
                            )
                            if not df.empty and df.shape[1] > 1:
                                return df
                        except Exception:
                            continue
        except Exception:
            continue
    return pd.DataFrame()


def criar_dataframe_completo(
    df_main: pd.DataFrame, df_ativos: pd.DataFrame
) -> pd.DataFrame:
    if df_main.empty:
        return pd.DataFrame()
    if df_ativos.empty:
        return df_main.copy()

    colunas_busca_ativos = [
        "LOGIN NETSALES",
        "LOGIN_NETSALES",
        "NETSALES",
        "NETSALES_LOGIN",
        "LOGIN",
        "MATRICULA",
        "MATRÍCULA",
        "RE",
        "ID",
        "CODIGO",
        "USUARIO",
        "TÉCNICO",
        "TECNICO",
        "NOME",
    ]
    colunas_busca_main = [
        "LOGIN NETSALES",
        "LOGIN_NETSALES",
        "NETSALES",
        "NETSALES_LOGIN",
        "LOGIN",
        "MATRICULA",
        "MATRÍCULA",
        "RE",
        "ID",
        "CODIGO",
        "USUARIO",
        "TÉCNICO",
        "TECNICO",
        "VENDEDOR",
        "COLABORADOR",
        "NOME",
    ]

    col_key_ativos = Utilitarios.buscar_coluna(df_ativos, colunas_busca_ativos)
    col_key_main = Utilitarios.buscar_coluna(df_main, colunas_busca_main)
    if not col_key_ativos or not col_key_main:
        return df_main.copy()

    main_df = df_main.copy()
    ativos = df_ativos.copy()
    main_df["_JOIN_KEY"] = main_df[col_key_main].astype(str).str.strip().str.upper()
    ativos["_JOIN_KEY"] = ativos[col_key_ativos].astype(str).str.strip().str.upper()

    merged = main_df.merge(
        ativos.drop_duplicates(subset=["_JOIN_KEY"]),
        on="_JOIN_KEY",
        how="left",
        suffixes=("", "_ativos"),
    ).drop(columns=["_JOIN_KEY"], errors="ignore")

    for col in list(merged.columns):
        if col.endswith("_ativos"):
            col_orig = col.replace("_ativos", "")
            if col_orig in merged.columns:
                merged[col_orig] = merged[col_orig].fillna(merged[col])
            else:
                merged[col_orig] = merged[col]
            merged.drop(columns=[col], inplace=True)
    return merged


# ====================================================
# 6. UTILITÁRIOS
# ====================================================
METAS_MENSAIS = (300.00, 350.00, 375.00, 400.00)
META_PRINCIPAL = 400.00


class Utilitarios:
    @staticmethod
    def buscar_coluna(df: pd.DataFrame, palavras_chave: List[str]) -> Optional[str]:
        cols_upper = {str(c).upper().strip(): str(c) for c in df.columns}
        for palavra in palavras_chave:
            key = palavra.upper().strip()
            if key in cols_upper:
                return cols_upper[key]
        for palavra in palavras_chave:
            key = palavra.upper().strip()
            for col_up, col_real in cols_upper.items():
                if key in col_up:
                    return col_real
        return None

    @staticmethod
    def formatar_numero(v: float, casas: int = 2) -> str:
        if pd.isna(v):
            return "0,00" if casas > 0 else "0"
        return f"{v:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")

    @staticmethod
    def formatar_posicao(valor: Any) -> str:
        try:
            v = int(valor)
            return {1: "🥇 1º", 2: "🥈 2º", 3: "🥉 3º"}.get(v, f"{v}º")
        except (ValueError, TypeError):
            return str(valor)

    @staticmethod
    def normalizar_pontos(series: pd.Series) -> pd.Series:
        if series is None or series.empty:
            return pd.Series(dtype=float)
        if pd.api.types.is_numeric_dtype(series):
            return series.fillna(0.0).astype(float)
        s = series.astype(str).str.strip().str.replace(r"[^\d.,\-]", "", regex=True)
        has_dot_and_comma = s.str.contains(r"\.") & s.str.contains(",")
        has_only_comma = s.str.contains(",") & ~s.str.contains(r"\.")
        s = s.where(
            ~has_dot_and_comma,
            s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
        )
        s = s.where(~has_only_comma, s.str.replace(",", ".", regex=False))
        return pd.to_numeric(s, errors="coerce").fillna(0.0).astype(float)

    @staticmethod
    def converter_data(series: pd.Series) -> pd.Series:
        if pd.api.types.is_datetime64_any_dtype(series):
            return series
        s = series.astype(str).str.strip()
        validos = s[~s.str.lower().isin(["", "nan", "nat", "none"])]
        if validos.empty:
            return pd.to_datetime(series, errors="coerce")
        if validos.str.match(r"^\d{4}-\d{1,2}-\d{1,2}").mean() > 0.6:
            return pd.to_datetime(s, errors="coerce")
        partes = validos.str.extract(r"^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})$")
        if partes.empty or partes[0].isna().all():
            return pd.to_datetime(s, errors="coerce", dayfirst=True)
        p1 = pd.to_numeric(partes[0], errors="coerce")
        p2 = pd.to_numeric(partes[1], errors="coerce")
        if (p1 > 12).any():
            return pd.to_datetime(s, errors="coerce", dayfirst=True)
        if (p2 > 12).any():
            return pd.to_datetime(s, errors="coerce", dayfirst=False)
        return pd.to_datetime(s, errors="coerce", dayfirst=True)

    @staticmethod
    def calcular_dias_uteis(
        df: pd.DataFrame, col_data: Optional[str]
    ) -> Tuple[int, int, Any, int]:
        data_referencia: Any = (
            pd.to_datetime(df[col_data].max()).date()
            if col_data and col_data in df.columns and pd.notna(df[col_data].max())
            else datetime.date.today()
        )
        ano, mes = data_referencia.year, data_referencia.month
        primeiro = datetime.date(ano, mes, 1)
        _, ult = calendar.monthrange(ano, mes)
        ultimo = datetime.date(ano, mes, ult)
        p_np = np.datetime64(primeiro)
        m_np = np.datetime64(data_referencia)
        u_np = np.datetime64(ultimo)
        wm = "1111110"
        total = int(np.busday_count(p_np, u_np + np.timedelta64(1, "D"), weekmask=wm))
        passados = int(
            np.busday_count(p_np, m_np + np.timedelta64(1, "D"), weekmask=wm)
        )
        brutos = max(0, total - passados)
        return brutos, max(1, brutos), data_referencia, passados

    @staticmethod
    def exportar_excel(abas: Dict[str, pd.DataFrame]) -> bytes:
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for nome, df in abas.items():
                sheet = str(nome)[:31] or "Aba"
                df.to_excel(writer, index=False, sheet_name=sheet)
                ws = writer.sheets[sheet]
                cor_cab = PatternFill("solid", fgColor="012869")
                cor_par = PatternFill("solid", fgColor="F8FAFC")
                cor_impar = PatternFill("solid", fgColor="FFFFFF")
                f_cab = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                borda = Border(
                    left=Side(style="thin", color="D9D9D9"),
                    right=Side(style="thin", color="D9D9D9"),
                    top=Side(style="thin", color="D9D9D9"),
                    bottom=Side(style="thin", color="D9D9D9"),
                )
                centro = Alignment(horizontal="center", vertical="center")
                for row in ws.iter_rows(
                    min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column
                ):
                    for cel in row:
                        cel.border = borda
                        if cel.row == 1:
                            cel.fill = cor_cab
                            cel.font = f_cab
                            cel.alignment = centro
                            continue
                        cel.fill = cor_par if cel.row % 2 == 0 else cor_impar
                for col in ws.columns:
                    letra = get_column_letter(col[0].column)
                    max_len = max(
                        (len(str(c.value)) for c in col if c.value is not None),
                        default=0,
                    )
                    ws.column_dimensions[letra].width = max(max_len + 3, 12)
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
        return output.getvalue()


class Configuracoes:
    VAZIOS = {"", "NAN", "NAT", "NONE", "NULL", "-", "N/A"}


class ProcessadorDeDados:
    @staticmethod
    def _normalizar(serie: pd.Series) -> pd.Series:
        s = serie.fillna("").astype(str).str.strip().str.upper()
        return s.where(~s.isin(Configuracoes.VAZIOS), "")

    @staticmethod
    def tratar_planos(df: pd.DataFrame) -> pd.DataFrame:
        if not {"PLANO TV", "PLANO INTERNET"}.issubset(df.columns):
            df["QTDE_CONSULTIVO"] = 0
            df["TIPO SERVIÇO"] = "Sem Tipo"
            return df

        internet_bruta = df["PLANO INTERNET"].astype(str).str.strip()
        partes = internet_bruta.str.split(".", n=1, expand=True)
        internet_limpa = ProcessadorDeDados._normalizar(partes[0])
        tv_embutida = (
            ProcessadorDeDados._normalizar(partes[1])
            if partes.shape[1] > 1
            else pd.Series("", index=df.index, dtype=str)
        )
        tv_original = ProcessadorDeDados._normalizar(df["PLANO TV"]).replace(
            "SERVIÇOS AVANÇADOS", "CLARO TV+ BOX"
        )
        tv_final = pd.Series(
            np.where(tv_original != "", tv_original, tv_embutida),
            index=df.index,
            dtype=str,
        )
        tem_tv = tv_final != ""
        tem_net = internet_limpa != ""
        df["QTDE_CONSULTIVO"] = tem_tv.astype(int) + tem_net.astype(int)
        cond = [tem_tv & tem_net, tem_tv & ~tem_net, ~tem_tv & tem_net]
        opts = [tv_final + " & " + internet_limpa, tv_final, internet_limpa]
        df["TIPO SERVIÇO"] = np.select(cond, opts, default="Sem Tipo")
        df["PLANO TV"] = tv_final
        df["PLANO INTERNET"] = internet_limpa
        return df

    @staticmethod
    def processar_quantidades(cons: pd.DataFrame) -> pd.DataFrame:
        col_obs = Utilitarios.buscar_coluna(
            cons, ["OBSERVACAO", "OBSERVAÇÃO", "OBS", "OBSERVACOES", "OBSERVAÇÕES"]
        )
        if col_obs and col_obs in cons.columns:
            cons["LISTA_PRODUTOS"] = (
                cons[col_obs].fillna("").astype(str).str.findall(r"\b\d{9,12}\b")
            )
            cons["QTDE_PRODUTOS"] = cons["LISTA_PRODUTOS"].str.len()
        else:
            cons["LISTA_PRODUTOS"] = [[] for _ in range(len(cons))]
            cons["QTDE_PRODUTOS"] = 0

        tipo_servico = (
            cons.get("TIPO SERVIÇO", pd.Series("", index=cons.index))
            .fillna("")
            .astype(str)
        )
        qtde_prod = cons["QTDE_PRODUTOS"].fillna(0).astype(int)
        is_combinado = tipo_servico.str.contains("&", case=False, regex=False)
        tem_tv = tipo_servico.str.contains("TV", case=False, regex=False)
        tem_virtua = tipo_servico.str.contains(r"MEGA|GIGA", case=False, regex=True)

        cons["QTDE_TV"] = np.where(
            is_combinado, tem_tv.astype(int), tem_tv.astype(int) * qtde_prod
        )
        cons["QTDE_VIRTUA"] = np.where(
            is_combinado, tem_virtua.astype(int), tem_virtua.astype(int) * qtde_prod
        )
        cons["QTDE_MESH"] = (qtde_prod - cons["QTDE_TV"] - cons["QTDE_VIRTUA"]).clip(
            lower=0
        )
        return cons

    @staticmethod
    def processar_completo(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or df.empty:
            return df
        df_out = df.copy()
        df_out = ProcessadorDeDados.tratar_planos(df_out)
        df_out = ProcessadorDeDados.processar_quantidades(df_out)
        return df_out


class ProcessamentoAdmin:
    @staticmethod
    def construir_ranking(
        df: pd.DataFrame,
        col_tec: str,
        col_pontos: str,
        dias_brutos: int,
        dias_seguros: int,
        dias_passados: int,
        col_equipe: Optional[str] = None,
        col_supervisor: Optional[str] = None,
        col_origem: Optional[str] = None,
    ) -> pd.DataFrame:
        group_cols = [col_tec]
        if col_equipe and col_equipe in df.columns and col_equipe not in group_cols:
            group_cols.append(col_equipe)
        if (
            col_supervisor
            and col_supervisor in df.columns
            and col_supervisor not in group_cols
        ):
            group_cols.append(col_supervisor)

        base = (
            df.groupby(group_cols, dropna=False)[col_pontos]
            .sum()
            .reset_index()
            .rename(columns={col_pontos: "Total Pontos"})
            .sort_values("Total Pontos", ascending=False)
            .reset_index(drop=True)
        )
        base.insert(0, "Posição", range(1, len(base) + 1))

        if col_origem and col_origem in df.columns:
            is_prod = (
                df[col_origem].astype(str).str.upper().str.contains("PROD", na=False)
            )
            os_count = df[is_prod].groupby(col_tec).size().reset_index(name="Qtd O.S.")
        else:
            os_count = df.groupby(col_tec).size().reset_index(name="Qtd O.S.")

        base = base.merge(os_count, on=col_tec, how="left").fillna({"Qtd O.S.": 0})

        if "Dias Trab Tecnico" in df.columns:
            dias_trab = (
                df.groupby(col_tec)["Dias Trab Tecnico"]
                .max()
                .fillna(dias_passados)
                .astype(int)
            )
        else:
            dias_trab = pd.Series(dias_passados, index=df[col_tec].unique())
        dias_trab = dias_trab.replace(0, 1)
        base["Dias Trab"] = (
            base[col_tec].map(dias_trab).fillna(dias_passados).astype(int)
        )
        base["Média/Dia"] = base["Total Pontos"] / base["Dias Trab"]
        base["Média/O.S."] = base["Total Pontos"] / base["Qtd O.S."].replace(0, 1)
        base["Projeção"] = base["Total Pontos"] + (base["Média/Dia"] * dias_brutos)
        return base

    @staticmethod
    def resumo_equipes(
        df_ranking: pd.DataFrame, col_equipe: str
    ) -> Optional[pd.DataFrame]:
        if col_equipe not in df_ranking.columns:
            return None
        return (
            df_ranking.groupby(col_equipe)
            .agg(
                Técnicos=("Posição", "count"),
                **{
                    "Total Pontos": ("Total Pontos", "sum"),
                    "Média Pontos": ("Total Pontos", "mean"),
                    "Projeção Média": ("Projeção", "mean"),
                    "Qtd O.S. Total": ("Qtd O.S.", "sum"),
                },
            )
            .sort_values("Total Pontos", ascending=False)
            .reset_index()
        )


# ====================================================
# 8. COMPONENTES VISUAIS
# ====================================================
_TEMAS_CARD = {
    "azul": {
        "fundo": "#EFF6FF",
        "texto": "#1E3A8A",
        "borda": "#3B82F6",
        "titulo": "#1E40AF",
    },
    "verde": {
        "fundo": "#F0FDF4",
        "texto": "#15803D",
        "borda": "#22C55E",
        "titulo": "#166534",
    },
    "laranja": {
        "fundo": "#FFF7ED",
        "texto": "#EA580C",
        "borda": "#F97316",
        "titulo": "#9A3412",
    },
    "cinza": {
        "fundo": "#F8FAFC",
        "texto": "#334155",
        "borda": "#94A3B8",
        "titulo": "#64748B",
    },
    "escuro": {
        "fundo": "#0F172A",
        "texto": "#F8FAFC",
        "borda": "#475569",
        "titulo": "#94A3B8",
    },
    "vermelho": {
        "fundo": "#FEF2F2",
        "texto": "#991B1B",
        "borda": "#EF4444",
        "titulo": "#7F1D1D",
    },
    "roxo": {
        "fundo": "#FAF5FF",
        "texto": "#6B21A8",
        "borda": "#A855F7",
        "titulo": "#7E22CE",
    },
}


def _card(
    titulo: str,
    valor: str,
    tema: str = "azul",
    sub: str = "",
    icone: str = "",
    tooltip: str = "",
) -> str:
    c = _TEMAS_CARD.get(tema, _TEMAS_CARD["azul"])
    sub_color = "#94A3B8" if tema == "escuro" else "#64748B"
    tip = f'<div class="tooltip-premium">{tooltip}</div>' if tooltip else ""
    return f"""
    <div class="card-admin" style="background:{c['fundo']};padding:20px;border-radius:10px;
         border-left:5px solid {c['borda']};box-shadow:0 4px 6px rgba(0,0,0,0.05);
         height:100%;display:flex;flex-direction:column;justify-content:center;
         transition:transform 0.2s,box-shadow 0.2s;"
         onmouseover="this.style.transform='translateY(-3px)';this.style.boxShadow='0 8px 15px rgba(0,0,0,0.1)';"
         onmouseout="this.style.transform='translateY(0)';this.style.boxShadow='0 4px 6px rgba(0,0,0,0.05)';">
        <p style="margin:0;font-size:13px;color:{c['titulo']};font-weight:700;">{icone} {titulo}</p>
        <h2 style="margin:5px 0 0;color:{c['texto']};font-weight:900;font-size:30px;">{valor}</h2>
        <p style="margin:5px 0 0;font-size:11px;color:{sub_color};font-weight:600;">{sub}</p>
        {tip}
    </div>"""


# ====================================================
# 9. INICIALIZAÇÃO E CARREGAMENTO
# ====================================================
aplicar_estilo()
aplicar_tema_claro()
injetar_css_admin()
render_sidebar_admin()

render_hero_totale_1(
    titulo="Painel Administrativo",
    subtitulo="Visão consolidada de produção e performance de toda a operação",
    icone="shield",
    badge="Área Restrita",
    usar_material=True,
)

loader = st.empty()
loader.markdown(
    """<div class="loading-totale">
        <span style="color:#012869;font-weight:700;">⚡ Carregando e unificando bases (Produção + Lista Ativos + Consultivos)...</span>
        <span style="color:#F37C04;font-weight:800;font-size:13px;">Aguarde</span>
    </div>""",
    unsafe_allow_html=True,
)

df_prod = carregar_base_producao()
df_ativos = carregar_lista_ativos()
df_consultivo_raw = carregar_base_consultivo()
loader.empty()

if df_prod.empty:
    render_insight(
        "Base de produção indisponível. Verifique o compartilhamento da planilha.",
        "critico",
    )
    st.stop()

df_raw = criar_dataframe_completo(df_prod, df_ativos)
if not df_consultivo_raw.empty and not df_ativos.empty:
    df_consultivo_raw = criar_dataframe_completo(df_consultivo_raw, df_ativos)

col_tec = Utilitarios.buscar_coluna(
    df_raw,
    [
        "LOGIN NETSALES",
        "LOGIN_NETSALES",
        "NETSALES",
        "NOME EQUIPE",
        "NOME_EQUIPE",
        "NOME TÉCNICO",
        "NOME_TECNICO",
        "TECNICO",
        "TÉCNICO",
        "NOME",
        "COLABORADOR",
        "VENDEDOR",
        "LOGIN",
        "USUARIO",
        "USER",
        "MATRICULA",
    ],
)
col_pontos = Utilitarios.buscar_coluna(
    df_raw, ["PONTOS", "PONTO", "PTS", "PONTUACAO", "PREVIA"]
)
col_data = Utilitarios.buscar_coluna(
    df_raw, ["DATA", "DATA AGENDAMENTO", "DATA CONCLUSÃO", "DATA_EXECUCAO", "DATE"]
)
col_equipe = Utilitarios.buscar_coluna(
    df_raw, ["EQUIPE", "NOME EQUIPE", "NOME_EQUIPE", "EQUIPE TÉCNICA"]
)
col_supervisor = Utilitarios.buscar_coluna(
    df_raw, ["SUPERVISOR", "SUPERVISOR EQUIPE", "GESTOR"]
)
col_origem = Utilitarios.buscar_coluna(
    df_raw, ["ORIGEM", "TIPO", "SISTEMA", "FONTE", "TIPO_OS"]
)
col_projeto = Utilitarios.buscar_coluna(df_raw, ["PROJETO", "CAMPANHA", "OPERAÇÃO"])

if not col_tec:
    st.error("🚨 Coluna de técnico não identificada na base de dados.")
    st.info(f"**Colunas detectadas:** \n\n {', '.join(df_raw.columns.tolist())}")
    st.stop()
if not col_pontos:
    st.error("🚨 Coluna de pontos não identificada na base de dados.")
    st.stop()

if col_tec:
    df_raw[col_tec] = df_raw[col_tec].astype(str).str.strip()
if col_equipe and col_equipe in df_raw.columns:
    df_raw[col_equipe] = df_raw[col_equipe].astype(str).str.strip()
if col_supervisor and col_supervisor in df_raw.columns:
    df_raw[col_supervisor] = df_raw[col_supervisor].astype(str).str.strip()
if col_projeto and col_projeto in df_raw.columns:
    df_raw[col_projeto] = df_raw[col_projeto].astype(str).str.strip()
if col_origem and col_origem in df_raw.columns:
    df_raw[col_origem] = df_raw[col_origem].astype(str).str.strip()

df_raw[col_pontos] = Utilitarios.normalizar_pontos(df_raw[col_pontos])
if col_data:
    df_raw[col_data] = Utilitarios.converter_data(df_raw[col_data]).dt.date

# ====================================================
# 10. PRÉ-PROCESSAMENTO CONSULTIVO
# ====================================================
df_cons_prep = pd.DataFrame()
col_cons_data = None
col_cons_tec = None
col_cons_nome = None
col_cons_produto = None
col_cons_qtd = None
col_cons_valor = None
col_cons_supervisor = None
col_cons_status = None
col_cons_base = None
col_cons_projeto = None

if df_consultivo_raw is not None and not df_consultivo_raw.empty:
    df_cons_prep = df_consultivo_raw.copy()

    col_cons_tec = Utilitarios.buscar_coluna(
        df_cons_prep,
        [
            "LOGIN NETSALES",
            "LOGIN_NETSALES",
            "NETSALES",
            "LOGIN",
            "USUARIO",
            "MATRICULA",
        ],
    )
    if not col_cons_tec:
        col_cons_tec = Utilitarios.buscar_coluna(
            df_cons_prep,
            ["TÉCNICO", "TECNICO", "VENDEDOR", "COLABORADOR", "NOME"],
        )
    col_cons_nome = Utilitarios.buscar_coluna(
        df_cons_prep,
        [
            "NOME TÉCNICO",
            "NOME_TECNICO",
            "NOME TECNICO",
            "TÉCNICO",
            "TECNICO",
            "NOME COLABORADOR",
            "COLABORADOR",
            "NOME VENDEDOR",
            "VENDEDOR",
            "NOME",
        ],
    )
    if col_cons_nome and col_cons_nome == col_cons_tec:
        col_cons_nome = None

    col_cons_data = Utilitarios.buscar_coluna(
        df_cons_prep,
        [
            "DATA",
            "DATA VENDA",
            "DATA_VENDA",
            "DATA CONSULTIVO",
            "DATA_CONSULTIVO",
            "DT",
        ],
    )
    col_cons_produto = Utilitarios.buscar_coluna(
        df_cons_prep,
        ["PRODUTO", "PRODUTOS", "ITEM", "SERVIÇO", "SERVICO", "OFERTA", "SKU"],
    )
    col_cons_qtd = Utilitarios.buscar_coluna(
        df_cons_prep, ["QUANTIDADE", "QTD", "QTDE", "QUANT"]
    )
    col_cons_valor = Utilitarios.buscar_coluna(
        df_cons_prep,
        ["VALOR", "VLR", "PREÇO", "PRECO", "RECEITA", "TICKET", "TOTAL"],
    )
    col_cons_supervisor = Utilitarios.buscar_coluna(
        df_cons_prep, ["SUPERVISOR", "GESTOR", "COORDENADOR"]
    )
    col_cons_status = Utilitarios.buscar_coluna(
        df_cons_prep, ["STATUS", "SITUAÇÃO", "SITUACAO", "RESULTADO"]
    )
    col_cons_base = Utilitarios.buscar_coluna(df_cons_prep, ["BASE"])
    col_cons_projeto = Utilitarios.buscar_coluna(
        df_cons_prep, ["PROJETO", "CAMPANHA", "OPERAÇÃO"]
    )

    if col_cons_data:
        df_cons_prep[col_cons_data] = Utilitarios.converter_data(
            df_cons_prep[col_cons_data]
        ).dt.date
    if col_cons_tec:
        df_cons_prep[col_cons_tec] = df_cons_prep[col_cons_tec].astype(str).str.strip()
    if col_cons_nome:
        df_cons_prep[col_cons_nome] = (
            df_cons_prep[col_cons_nome].astype(str).str.strip().str.title()
        )
    if col_cons_produto:
        df_cons_prep[col_cons_produto] = (
            df_cons_prep[col_cons_produto].astype(str).str.strip().str.upper()
        )
    if col_cons_supervisor:
        df_cons_prep[col_cons_supervisor] = (
            df_cons_prep[col_cons_supervisor].astype(str).str.strip()
        )
    if col_cons_status:
        df_cons_prep[col_cons_status] = (
            df_cons_prep[col_cons_status].astype(str).str.strip().str.upper()
        )
    if col_cons_qtd:
        df_cons_prep[col_cons_qtd] = Utilitarios.normalizar_pontos(
            df_cons_prep[col_cons_qtd]
        )
    if col_cons_valor:
        df_cons_prep[col_cons_valor] = Utilitarios.normalizar_pontos(
            df_cons_prep[col_cons_valor]
        )
    if col_cons_base:
        df_cons_prep[col_cons_base] = (
            df_cons_prep[col_cons_base].astype(str).str.strip().str.upper()
        )
    if col_cons_projeto:
        df_cons_prep[col_cons_projeto] = (
            df_cons_prep[col_cons_projeto].astype(str).str.strip().str.upper()
        )

    df_cons_prep = ProcessadorDeDados.processar_completo(df_cons_prep)


# ====================================================
# 11. FILTROS GLOBAIS E KPIs (Visão Geral em cima, Filtros embaixo)
# ====================================================

# ── Placeholder para renderizar a Visão Geral ANTES dos Filtros ──
placeholder_visao_geral = st.container()
st.divider()

# ── 🎯 FILTROS GLOBAIS UNIFICADOS (Produção & Consultivo) ──
with st.container(border=True):
    render_section_header("🎯", "Filtros Globais")
    f1, f2, f3, f4 = st.columns([1, 1, 1.2, 1])

    # 1. Determinação de limites temporais baseados na Produção (Âncora temporal)
    max_d = datetime.date.today()
    min_d = max_d - timedelta(days=30)  # Limite padrão seguro caso esteja vazia

    if col_data and not df_raw[col_data].dropna().empty:
        min_d = df_raw[col_data].dropna().min()
        max_d = max(df_raw[col_data].dropna().max(), datetime.date.today())

    with f1:
        d_ini = st.date_input(
            "📅 De",
            value=min_d,
            min_value=min_d,
            max_value=max_d,
            format="DD/MM/YYYY",
            key="filtro_data_ini",
        )
    with f2:
        d_fim = st.date_input(
            "📅 Até",
            value=max_d,
            min_value=min_d,
            max_value=max_d,
            format="DD/MM/YYYY",
            key="filtro_data_fim",
        )

    with f3:
        projeto_sel = "Todos"
        if col_projeto and col_projeto in df_raw.columns:
            projetos = sorted(df_raw[col_projeto].dropna().unique().tolist())
            opcoes_projetos = ["Todos", *projetos]
            target_proj = "NET-ABCDM"
            index_padrao = 0
            for idx, op in enumerate(opcoes_projetos):
                if str(op).strip().upper() == target_proj:
                    index_padrao = idx
                    break
            projeto_sel = st.selectbox(
                "📁 Projetos",
                opcoes_projetos,
                index=index_padrao,
                key="filtro_projeto",
            )

    with f4:
        quick = st.selectbox(
            "⚡ Atalho",
            ["Todos", "Hoje", "Últimos 7 dias", "Mês Atual"],
            key="filtro_atalho",
        )

    # 3. Inicialização das Máscaras Booleanas de Filtragem
    mask_prod = pd.Series(True, index=df_raw.index)
    mask_cons = (
        pd.Series(True, index=df_cons_prep.index) if not df_cons_prep.empty else None
    )

    # 4. Aplicação de Período Manual
    if d_ini <= d_fim:
        if col_data:
            mask_prod &= (df_raw[col_data] >= d_ini) & (df_raw[col_data] <= d_fim)
        if mask_cons is not None and col_cons_data:
            mask_cons &= (df_cons_prep[col_cons_data] >= d_ini) & (
                df_cons_prep[col_cons_data] <= d_fim
            )
    else:
        render_insight("Data inicial não pode ser posterior à data final.", "alerta")

    # 5. Aplicação de Atalhos Rápidos (Sincronização entre as bases)
    if quick != "Todos":
        if quick == "Hoje":
            d_ref_ini, d_ref_fim = max_d, max_d
        elif quick == "Últimos 7 dias":
            d_ref_ini, d_ref_fim = max_d - timedelta(days=6), max_d
        elif quick == "Mês Atual":
            d_ref_ini, d_ref_fim = max_d.replace(day=1), max_d

        if col_data:
            mask_prod &= (df_raw[col_data] >= d_ref_ini) & (
                df_raw[col_data] <= d_ref_fim
            )
        if mask_cons is not None and col_cons_data:
            mask_cons &= (df_cons_prep[col_cons_data] >= d_ref_ini) & (
                df_cons_prep[col_cons_data] <= d_ref_fim
            )

    # 6. Aplicação de Projeto
    if projeto_sel != "Todos":
        if col_projeto:
            mask_prod &= df_raw[col_projeto] == projeto_sel
        if mask_cons is not None and col_cons_projeto:
            mask_cons &= (
                df_cons_prep[col_cons_projeto].astype(str).str.upper()
                == str(projeto_sel).upper()
            )

# ── APLICAÇÃO FINAL E CRIAÇÃO DAS BASES FILTRADAS ──
df_filtered = df_raw[mask_prod].copy()

df_cons_filtered = pd.DataFrame()
if mask_cons is not None:
    df_cons_filtered = df_cons_prep[mask_cons].copy()

# Interrupção de segurança caso a Produção retorne vazia
if df_filtered.empty:
    with placeholder_visao_geral:
        st.info("💡 Nenhum dado de produção encontrado para os filtros selecionados.")
    st.stop()

# ── PROCESSAMENTO DE DIAS E MÉTRICAS DA PRODUÇÃO ──
dias_brutos, dias_seguros, data_ref, dias_passados = Utilitarios.calcular_dias_uteis(
    df_filtered, col_data
)
total_pontos = float(df_filtered[col_pontos].sum())
total_tecnicos = df_filtered[col_tec].nunique()
pts_prod, pts_gpon, total_os = 0.0, 0.0, 0
if col_origem:
    origem = df_filtered[col_origem].astype(str).str.upper()
    is_prod = origem.str.contains("PROD", na=False)
    pts_prod = float(df_filtered.loc[is_prod, col_pontos].sum())
    pts_gpon = float(df_filtered.loc[~is_prod, col_pontos].sum())
    total_os = int(is_prod.sum())
else:
    pts_gpon = total_pontos
    total_os = len(df_filtered)

tec_acima_400, tec_criticos = 0, 0
if total_tecnicos > 0:
    pts_por_tec = df_filtered.groupby(col_tec)[col_pontos].sum()
    tec_acima_400 = int((pts_por_tec >= 400.00).sum())
    tec_criticos = int(((pts_por_tec >= 275.00) & (pts_por_tec < 300.00)).sum())

# ── RENDERIZAÇÃO DA VISÃO GERAL NO PLACEHOLDER (Acima dos Filtros) ──
with placeholder_visao_geral:
    render_section_header("📊", "Visão Geral da Operação")

    st.markdown(
        "<p style='color:#012869;font-weight:800;font-size:13px;margin:0 0 8px;letter-spacing:1px;'>🎯 PRODUÇÃO</p>",
        unsafe_allow_html=True,
    )
    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        st.markdown(
            _card(
                "Total de Pontos",
                Utilitarios.formatar_numero(total_pontos, casas=2),
                "azul",
                f"Prod: {Utilitarios.formatar_numero(pts_prod, casas=2)} | Gpon: {Utilitarios.formatar_numero(pts_gpon, casas=2)}",
                "🎯",
            ),
            unsafe_allow_html=True,
        )
    with k2:
        st.markdown(
            _card(
                "O.S. Realizadas (Prod)",
                str(total_os),
                "cinza",
                "Apenas atendimentos de Produção",
                "📋",
            ),
            unsafe_allow_html=True,
        )
    with k3:
        st.markdown(
            _card(
                "Técnicos Ativos",
                str(total_tecnicos),
                "verde",
                f"{dias_passados} dias úteis passados",
                "👷",
            ),
            unsafe_allow_html=True,
        )
    with k4:
        st.markdown(
            _card(
                "Elite (≥ 400,00 pts)",
                f"{tec_acima_400}/{total_tecnicos}",
                "escuro",
                "Técnicos que já bateram a meta principal",
                "🏆",
            ),
            unsafe_allow_html=True,
        )
    with k5:
        st.markdown(
            _card(
                "Próximo da Meta (275-300)",
                str(tec_criticos),
                "vermelho" if tec_criticos > 0 else "verde",
                f"{(tec_criticos / max(total_tecnicos, 1) * 100):.1f}% da equipe",
                "⚠️",
            ),
            unsafe_allow_html=True,
        )

    st.write("")
    st.markdown(
        "<p style='color:#7E22CE;font-weight:800;font-size:13px;margin:0 0 8px;letter-spacing:1px;'>💼 CONSULTIVO</p>",
        unsafe_allow_html=True,
    )

    if df_cons_prep.empty:
        st.info(
            "💡 Base de consultivos indisponível. Faça upload na aba **Visão Consultivo** para integrar."
        )
    elif df_cons_filtered.empty:
        st.info("💡 Nenhum consultivo encontrado para os filtros globais aplicados.")
    else:
        total_cons_registros = len(df_cons_filtered)
        total_qtde_consultivo = (
            int(df_cons_filtered["QTDE_CONSULTIVO"].sum())
            if "QTDE_CONSULTIVO" in df_cons_filtered.columns
            else total_cons_registros
        )
        total_qtde_produtos = (
            int(df_cons_filtered["QTDE_PRODUTOS"].sum())
            if "QTDE_PRODUTOS" in df_cons_filtered.columns
            else 0
        )
        total_vend = df_cons_filtered[col_cons_tec].nunique() if col_cons_tec else 0
        total_receita = (
            float(df_cons_filtered[col_cons_valor].sum()) if col_cons_valor else 0.0
        )
        ticket = (
            total_receita / total_cons_registros
            if total_cons_registros and col_cons_valor
            else 0.0
        )

        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.markdown(
                _card(
                    "Registros de Vendas",
                    Utilitarios.formatar_numero(total_cons_registros, casas=0),
                    "roxo",
                    "Total de linhas na base",
                    "💼",
                ),
                unsafe_allow_html=True,
            )
        with c2:
            st.markdown(
                _card(
                    "Qtde Consultivos",
                    Utilitarios.formatar_numero(total_qtde_consultivo, casas=0),
                    "azul",
                    "Serviços (TV + NET) contratados",
                    "🔗",
                ),
                unsafe_allow_html=True,
            )
        with c3:
            st.markdown(
                _card(
                    "Qtde Produtos",
                    Utilitarios.formatar_numero(total_qtde_produtos, casas=0),
                    "laranja",
                    "Produtos identificados na OBS",
                    "📦",
                ),
                unsafe_allow_html=True,
            )
        with c4:
            st.markdown(
                _card(
                    "Vendedores Ativos",
                    str(total_vend),
                    "verde",
                    "Consultivos executados",
                    "👥",
                ),
                unsafe_allow_html=True,
            )
        with c5:
            if col_cons_valor and total_receita > 0:
                st.markdown(
                    _card(
                        "Receita Consultiva",
                        f"R$ {Utilitarios.formatar_numero(total_receita, casas=2)}",
                        "escuro",
                        f"Ticket médio: R$ {Utilitarios.formatar_numero(ticket, casas=2)}",
                        "💰",
                    ),
                    unsafe_allow_html=True,
                )
            else:
                media_cons = total_qtde_consultivo / max(total_vend, 1)
                st.markdown(
                    _card(
                        "Média/Vendedor",
                        f"{media_cons:.2f}",
                        "escuro",
                        "Consultivos por vendedor",
                        "📈",
                    ),
                    unsafe_allow_html=True,
                )

# ── Gerar ranking base global para as abas subsequentes ──
df_ranking = ProcessamentoAdmin.construir_ranking(
    df_filtered,
    col_tec,
    col_pontos,
    dias_brutos,
    dias_seguros,
    dias_passados,
    col_equipe=col_equipe,
    col_supervisor=col_supervisor,
    col_origem=col_origem,
)
df_ranking["Status"] = np.select(
    [
        df_ranking["Total Pontos"] >= 400.00,
        df_ranking["Total Pontos"] >= 300.00,
        df_ranking["Total Pontos"] >= 275.00,
    ],
    ["🏆 Elite", "✅ Meta", "🎯 Próximo da Meta"],
    default="🔴 Crítico",
)

st.divider()

# ====================================================
# 12. TABS PRINCIPAIS
# ====================================================
(
    tab_visao_geral,
    tab_equipes,
    tab_graficos,
    tab_diagnostico,
    tab_consultivo,
    tab_logins,
) = st.tabs(
    [
        "🏆 Produção",
        "👥 Por Equipe",
        "📈 Gráficos",
        "📊 Diagnóstico",
        "💼 Visão Consultivo",
        "🔐 Logins",
    ]
)

# ====================================================
# ── TAB 1: RANKING DE PRODUÇÃO CONSOLIDADA ──
# ====================================================
with tab_visao_geral:
    render_section_header("🏆", "Produção Consolidada por Técnico")

    r1, r2, r3 = st.columns(3)
    with r1:
        proj_media = df_ranking["Projeção"].mean()
        st.markdown(
            _card(
                "Projeção Média Equipe",
                Utilitarios.formatar_numero(proj_media, casas=2),
                "escuro",
                "Fim do mês estimado",
                "📈",
            ),
            unsafe_allow_html=True,
        )
    with r2:
        top1 = df_ranking.iloc[0][col_tec] if not df_ranking.empty else "—"
        top1_pts = df_ranking.iloc[0]["Total Pontos"] if not df_ranking.empty else 0.00
        st.markdown(
            _card(
                "Técnico com Maior Pontuação",
                str(top1).title(),
                "laranja",
                f"{Utilitarios.formatar_numero(top1_pts, casas=2)} pts",
                "🥇",
            ),
            unsafe_allow_html=True,
        )
    with r3:
        st.markdown(
            _card(
                "Dias Úteis Restantes",
                str(dias_brutos),
                "cinza",
                f"Ref: {pd.Timestamp(data_ref).strftime('%d/%m/%Y')}",
                "📆",
            ),
            unsafe_allow_html=True,
        )

    st.write("")
    colunas_ranking = ["Posição", col_tec]
    if col_supervisor and col_supervisor in df_ranking.columns:
        colunas_ranking.append(col_supervisor)
    colunas_ranking += [
        "Status",
        "Qtd O.S.",
        "Total Pontos",
        "Dias Trab",
        "Média/Dia",
        "Média/O.S.",
        "Projeção",
    ]

    df_rank_display = df_ranking[colunas_ranking].copy()
    df_rank_display["Posição"] = df_rank_display["Posição"].apply(
        Utilitarios.formatar_posicao
    )
    col_config = {
        "Posição": st.column_config.TextColumn("Pos."),
        "Dias Trab": st.column_config.NumberColumn("Dias Trab.", format="%d"),
        col_tec: st.column_config.TextColumn("Nome Técnico"),
        col_supervisor: st.column_config.TextColumn("Supervisor"),
        "Status": st.column_config.TextColumn("Status"),
        "Qtd O.S.": st.column_config.NumberColumn("📋 O.S.", format="%d"),
        "Total Pontos": st.column_config.NumberColumn("🎯 Pontos", format="%.2f"),
        "Média/Dia": st.column_config.NumberColumn("⚡ Méd/Dia", format="%.2f"),
        "Média/O.S.": st.column_config.NumberColumn("📊 Méd/O.S.", format="%.2f"),
        "Projeção": st.column_config.NumberColumn("📈 Projeção", format="%.2f"),
    }
    st.dataframe(
        df_rank_display,
        use_container_width=True,
        hide_index=True,
        height=500,
        column_config=col_config,
    )

# ====================================================
# ── TAB 2: POR EQUIPE ──
# ====================================================
with tab_equipes:
    render_section_header("👥", "Performance por Equipe")
    if df_ranking.empty:
        st.info("💡 Carregue dados na aba de Produção primeiro.")
    elif col_equipe and col_equipe in df_ranking.columns:
        df_eq = ProcessamentoAdmin.resumo_equipes(df_ranking, col_equipe)
        if df_eq is not None and not df_eq.empty:
            st.dataframe(
                df_eq,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Total Pontos": st.column_config.NumberColumn(
                        "🎯 Total Pontos", format="%.2f"
                    ),
                    "Média Pontos": st.column_config.NumberColumn(
                        "📊 Média/Téc", format="%.2f"
                    ),
                    "Projeção Média": st.column_config.NumberColumn(
                        "📈 Projeção Méd", format="%.2f"
                    ),
                    "Qtd O.S. Total": st.column_config.NumberColumn(
                        "📋 O.S.", format="%d"
                    ),
                    "Técnicos": st.column_config.NumberColumn(
                        "👷 Técnicos", format="%d"
                    ),
                },
            )
            fig_eq = px.bar(
                df_eq.sort_values("Total Pontos"),
                x="Total Pontos",
                y=col_equipe,
                orientation="h",
                color="Média Pontos",
                color_continuous_scale=["#EF4444", "#F97316", "#22C55E", "#012869"],
                text_auto=True,
                title="Pontos Totais por Equipe",
            )
            fig_eq.update_traces(texttemplate="%{x:.2f}")
            fig_eq.update_layout(
                margin=dict(l=0, r=20, t=40, b=0),
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(
                fig_eq, use_container_width=True, config={"displayModeBar": False}
            )
        else:
            st.info("💡 Sem dados de equipe disponíveis.")
    else:
        st.info("💡 Coluna de equipe não identificada na base de dados.")

# ====================================================
# ── TAB 3: GRÁFICOS ──
# ====================================================
with tab_graficos:
    render_section_header("📈", "Análises Gráficas")
    if df_ranking.empty:
        st.info("💡 Dados do ranking não calculados.")
    else:
        g1, g2 = st.columns(2)
        with g1:
            st.subheader("🏆 Top 15 Técnicos")
            top15 = df_ranking.head(15).copy()
            fig_top = go.Figure(
                go.Bar(
                    x=top15["Total Pontos"],
                    y=top15[col_tec].astype(str).str.title(),
                    orientation="h",
                    marker_color=np.where(
                        top15["Total Pontos"] >= 400.00,
                        "#012869",
                        np.where(
                            top15["Total Pontos"] >= 300.00,
                            "#22C55E",
                            np.where(
                                top15["Total Pontos"] >= 275.00, "#F97316", "#EF4444"
                            ),
                        ),
                    ),
                    text=top15["Total Pontos"].apply(lambda x: f"{x:.2f}"),
                    textposition="outside",
                )
            )
            fig_top.update_layout(
                margin=dict(l=0, r=60, t=10, b=0),
                yaxis=dict(autorange="reversed"),
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.05)"),
            )
            st.plotly_chart(
                fig_top, use_container_width=True, config={"displayModeBar": False}
            )
        with g2:
            st.subheader("📊 Distribuição de Pontos")
            fig_hist = go.Figure(
                go.Histogram(
                    x=df_ranking["Total Pontos"],
                    nbinsx=20,
                    marker_color="#012869",
                    opacity=0.8,
                )
            )
            for meta, cor in [
                (300.00, "#22C55E"),
                (350.00, "#F97316"),
                (400.00, "#012869"),
            ]:
                fig_hist.add_vline(
                    x=meta,
                    line_dash="dash",
                    line_color=cor,
                    annotation_text=f"Meta {meta:.1f}",
                    annotation_position="top",
                )
            fig_hist.update_layout(
                margin=dict(l=0, r=20, t=10, b=0),
                xaxis_title="Pontos",
                yaxis_title="Técnicos",
                plot_bgcolor="rgba(0,0,0,0)",
                paper_bgcolor="rgba(0,0,0,0)",
            )
            st.plotly_chart(
                fig_hist, use_container_width=True, config={"displayModeBar": False}
            )

# ====================================================
# ── TAB 4: DIAGNÓSTICO ──
# ====================================================
with tab_diagnostico:
    render_section_header("📊", "Painel de Diagnóstico")
    if df_ranking.empty:
        st.info("💡 Dados do ranking indisponíveis.")
    else:
        if col_supervisor and col_supervisor in df_ranking.columns:
            st.subheader("👥 Matriz de Liderança e Eficiência")
            df_supervisores = (
                df_ranking.groupby(col_supervisor)
                .agg(
                    Tecnicos=("Posição", "count"),
                    Total_Pontos=("Total Pontos", "sum"),
                    Media_Pontos=("Total Pontos", "mean"),
                    OS_Totais=("Qtd O.S.", "sum"),
                )
                .reset_index()
            )
            df_supervisores["Pontos por O.S."] = df_supervisores[
                "Total_Pontos"
            ] / df_supervisores["OS_Totais"].replace(0, 1)
            df_supervisores = df_supervisores.sort_values(
                "Total_Pontos", ascending=False
            )

            sc1, sc2 = st.columns([1.5, 1])
            with sc1:
                st.dataframe(
                    df_supervisores,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        col_supervisor: st.column_config.TextColumn("Supervisor"),
                        "Tecnicos": st.column_config.NumberColumn(
                            "👷 Técnicos Ativos", format="%d"
                        ),
                        "Total_Pontos": st.column_config.NumberColumn(
                            "🎯 Total Pontos", format="%.2f"
                        ),
                        "Media_Pontos": st.column_config.NumberColumn(
                            "📊 Média/Téc", format="%.2f"
                        ),
                        "OS_Totais": st.column_config.NumberColumn(
                            "📋 O.S. Realizadas", format="%d"
                        ),
                        "Pontos por O.S.": st.column_config.NumberColumn(
                            "⚡ Eficiência (Pts/O.S.)", format="%.2f"
                        ),
                    },
                )
            with sc2:
                fig_eff = px.bar(
                    df_supervisores,
                    x="Pontos por O.S.",
                    y=col_supervisor,
                    orientation="h",
                    color="Pontos por O.S.",
                    color_continuous_scale="Purples",
                    title="Eficiência (Pontos por O.S.)",
                    text_auto=True,
                )
                fig_eff.update_traces(texttemplate="%{x:.2f}")
                fig_eff.update_layout(
                    margin=dict(l=0, r=10, t=30, b=0),
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                )
                st.plotly_chart(
                    fig_eff, use_container_width=True, config={"displayModeBar": False}
                )

        st.divider()
        st.subheader("🚨 Plano de Recuperação de Performance")
        st.info("🎯 Alvo: Técnicos com menos de 300,00 pontos acumulados no mês.")
        criticos_reais = df_ranking[df_ranking["Total Pontos"] < 300.00].copy()
        if not criticos_reais.empty:
            criticos_reais["Falta para Meta"] = 300.00 - criticos_reais["Total Pontos"]
            cols_criticas = ["Posição", col_tec]
            if col_supervisor and col_supervisor in criticos_reais.columns:
                cols_criticas.append(col_supervisor)
            cols_criticas += [
                "Total Pontos",
                "Falta para Meta",
                "Média/Dia",
                "Projeção",
            ]
            st.dataframe(
                criticos_reais[cols_criticas],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Posição": st.column_config.TextColumn("Pos."),
                    "Total Pontos": st.column_config.NumberColumn(
                        "🎯 Pontos Atuais", format="%.2f"
                    ),
                    "Falta para Meta": st.column_config.NumberColumn(
                        "⚠️ Falta p/ Meta", format="%.2f"
                    ),
                    "Média/Dia": st.column_config.NumberColumn(
                        "⚡ Média/Dia", format="%.2f"
                    ),
                    "Projeção": st.column_config.NumberColumn(
                        "📈 Projeção Fim de Mês", format="%.2f"
                    ),
                    **(
                        {col_tec: st.column_config.TextColumn("Técnico em Risco")}
                        if col_tec
                        else {}
                    ),
                    **(
                        {
                            col_supervisor: st.column_config.TextColumn(
                                "Supervisor Responsável"
                            )
                        }
                        if col_supervisor
                        else {}
                    ),
                },
            )
        else:
            st.success("🎉 Incrível! 100% da equipe atingiu 300,00 pontos!")

        st.divider()
        st.subheader("🔍 Ficha de Diagnóstico Individual")
        lista_tecnicos = sorted(df_ranking[col_tec].dropna().unique().tolist())
        if lista_tecnicos:
            tec_selecionado = st.selectbox("Selecione um técnico", lista_tecnicos)
            ficha_tec = df_ranking[df_ranking[col_tec] == tec_selecionado].iloc[0]
            c_diag1, c_diag2, c_diag3, c_diag4 = st.columns(4)
            with c_diag1:
                st.metric("🏆 Posição", f"#{ficha_tec['Posição']}")
            with c_diag2:
                st.metric("🎯 Pontuação Total", f"{ficha_tec['Total Pontos']:.2f} pts")
            with c_diag3:
                st.metric(
                    "📋 Produtividade Média", f"{ficha_tec['Média/O.S.']:.2f} pts/O.S."
                )
            with c_diag4:
                st.metric(
                    "📈 Estimativa Fim de Mês", f"{ficha_tec['Projeção']:.2f} pts"
                )

            pts_atual = float(ficha_tec["Total Pontos"])
            pct_meta = min(1.00, pts_atual / 400.00)
            st.write("")
            st.markdown(
                f"**Progresso para a Meta Principal (400,00):** {pct_meta*100:.1f}%"
            )
            st.progress(pct_meta)

            if col_data and col_data in df_filtered.columns:
                df_tec_data = df_filtered[
                    df_filtered[col_tec] == tec_selecionado
                ].copy()
                df_tec_data_agrupado = (
                    df_tec_data.groupby(col_data)[col_pontos]
                    .sum()
                    .reset_index()
                    .sort_values(col_data)
                )
                fig_evol = px.line(
                    df_tec_data_agrupado,
                    x=col_data,
                    y=col_pontos,
                    title=f"Evolução Diária — {tec_selecionado}",
                    labels={col_data: "Data", col_pontos: "Pontos do Dia"},
                    markers=True,
                )
                fig_evol.update_traces(line_color="#F97316", marker=dict(size=8))
                fig_evol.update_layout(
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.05)"),
                )
                st.plotly_chart(
                    fig_evol, use_container_width=True, config={"displayModeBar": False}
                )
        else:
            st.info("Sem técnicos disponíveis.")

# ====================================================
# ── TAB 5: CONSULTIVO ──
# ====================================================
with tab_consultivo:
    render_section_header("💼", "Vendas Consultivas & Produtos")

    if df_consultivo_raw is None or df_consultivo_raw.empty:
        st.warning("⚠️ Não foi possível baixar a base do Google Drive automaticamente.")
        st.info("💡 **Solução:** Carregue o arquivo abaixo:")
        up_file = st.file_uploader(
            "Upload da Planilha de Consultivos",
            type=["xlsx", "xls", "csv"],
            key="uploader_consultivo",
        )
        if up_file is not None:
            try:
                if up_file.name.endswith((".xlsx", ".xls")):
                    excel_file = pd.read_excel(
                        up_file, sheet_name=None, engine="openpyxl"
                    )
                    dfs = []
                    for s_name, df_s in excel_file.items():
                        if not df_s.empty:
                            df_s["_ABA_ORIGEM"] = s_name
                            dfs.append(df_s)
                    df_uploaded = (
                        pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()
                    )
                else:
                    df_uploaded = pd.read_csv(up_file)

                if not df_uploaded.empty and not df_ativos.empty:
                    df_consultivo_raw = criar_dataframe_completo(df_uploaded, df_ativos)
                else:
                    df_consultivo_raw = df_uploaded
            except Exception as e:
                st.error(f"Erro ao processar o arquivo: {e}")

    # Aplicação do baseline filtrado globalmente
    if not df_cons_prep.empty:
        df_cons = (
            df_cons_filtered.copy()
            if not df_cons_filtered.empty
            else df_cons_prep.copy()
        )

        with st.container(border=True):
            st.markdown("##### 🎯 Filtros Adicionais de Consultivos (Tab Local)")
            fc1, fc2 = st.columns(2)
            mask_cons_local = pd.Series(True, index=df_cons.index)

            with fc1:
                if col_cons_supervisor:
                    supervisores_cons = ["Todos"] + sorted(
                        df_cons[col_cons_supervisor].dropna().unique().tolist()
                    )
                    sup_sel = st.selectbox(
                        "👤 Supervisor", supervisores_cons, key="sup_cons"
                    )
                    if sup_sel != "Todos":
                        mask_cons_local &= df_cons[col_cons_supervisor] == sup_sel
                else:
                    st.write("")

            with fc2:
                if col_cons_base:
                    bases_cons = ["Todos"] + sorted(
                        df_cons[col_cons_base].dropna().unique().tolist()
                    )
                    base_sel = st.selectbox("🏢 Base", bases_cons, key="base_cons")
                    if base_sel != "Todos":
                        mask_cons_local &= df_cons[col_cons_base] == base_sel

        df_cons_f = df_cons[mask_cons_local].copy()

        if df_cons_f.empty:
            st.info(
                "💡 Nenhum registro consultivo encontrado com os filtros aplicados."
            )
        else:
            total_registros = len(df_cons_f)
            total_vendedores = df_cons_f[col_cons_tec].nunique() if col_cons_tec else 0
            total_qtde_cons = (
                int(df_cons_f["QTDE_CONSULTIVO"].sum())
                if "QTDE_CONSULTIVO" in df_cons_f.columns
                else total_registros
            )
            total_qtde_prods = (
                int(df_cons_f["QTDE_PRODUTOS"].sum())
                if "QTDE_PRODUTOS" in df_cons_f.columns
                else 0
            )
            total_tv = (
                int(df_cons_f["QTDE_TV"].sum()) if "QTDE_TV" in df_cons_f.columns else 0
            )
            total_virtua = (
                int(df_cons_f["QTDE_VIRTUA"].sum())
                if "QTDE_VIRTUA" in df_cons_f.columns
                else 0
            )
            total_mesh = (
                int(df_cons_f["QTDE_MESH"].sum())
                if "QTDE_MESH" in df_cons_f.columns
                else 0
            )
            total_valor = (
                float(df_cons_f[col_cons_valor].sum()) if col_cons_valor else 0.0
            )
            ticket_medio = (
                (total_valor / total_registros)
                if (col_cons_valor and total_registros)
                else 0.0
            )

            st.write("")
            kc1, kc2, kc3, kc4, kc5 = st.columns(5)
            with kc1:
                st.markdown(
                    _card(
                        "Registros",
                        Utilitarios.formatar_numero(total_registros, casas=0),
                        "azul",
                        "Total de linhas na base",
                        "💼",
                    ),
                    unsafe_allow_html=True,
                )
            with kc2:
                st.markdown(
                    _card(
                        "Qtde Consultivos",
                        Utilitarios.formatar_numero(total_qtde_cons, casas=0),
                        "roxo",
                        "Serviços TV + NET",
                        "🔗",
                    ),
                    unsafe_allow_html=True,
                )
            with kc3:
                st.markdown(
                    _card(
                        "Qtde Produtos",
                        Utilitarios.formatar_numero(total_qtde_prods, casas=0),
                        "laranja",
                        "Extraídos da OBS",
                        "📦",
                    ),
                    unsafe_allow_html=True,
                )
            with kc4:
                st.markdown(
                    _card(
                        "Vendedores Ativos",
                        str(total_vendedores),
                        "verde",
                        "Diferentes atuantes",
                        "👥",
                    ),
                    unsafe_allow_html=True,
                )
            with kc5:
                if col_cons_valor and total_valor > 0:
                    st.markdown(
                        _card(
                            "Receita Total",
                            f"R$ {Utilitarios.formatar_numero(total_valor, casas=2)}",
                            "escuro",
                            f"Ticket médio: R$ {Utilitarios.formatar_numero(ticket_medio, casas=2)}",
                            "💰",
                        ),
                        unsafe_allow_html=True,
                    )
                else:
                    media = total_qtde_cons / max(total_vendedores, 1)
                    st.markdown(
                        _card(
                            "Média/Vendedor",
                            f"{media:.2f}",
                            "escuro",
                            "Consultivos por vendedor",
                            "📈",
                        ),
                        unsafe_allow_html=True,
                    )

            st.write("")
            b1, b2, b3 = st.columns(3)
            with b1:
                st.markdown(
                    _card(
                        "📺 TV",
                        Utilitarios.formatar_numero(total_tv, casas=0),
                        "azul",
                        "Unidades de TV vendidas",
                        "📺",
                    ),
                    unsafe_allow_html=True,
                )
            with b2:
                st.markdown(
                    _card(
                        "🌐 Virtua (Mega/Giga)",
                        Utilitarios.formatar_numero(total_virtua, casas=0),
                        "verde",
                        "Planos de Internet Virtua",
                        "🌐",
                    ),
                    unsafe_allow_html=True,
                )
            with b3:
                st.markdown(
                    _card(
                        "📡 MESH & Outros",
                        Utilitarios.formatar_numero(total_mesh, casas=0),
                        "laranja",
                        "Roteadores e demais produtos",
                        "📡",
                    ),
                    unsafe_allow_html=True,
                )

            st.divider()

            df_vend_rank = pd.DataFrame()
            if col_cons_tec:
                st.subheader("🏆 Ranking de Vendedores Consultivos")

                group_keys = [col_cons_tec]
                if col_cons_nome and col_cons_nome in df_cons_f.columns:
                    group_keys.append(col_cons_nome)

                agrup_vend = {"Registros": (col_cons_tec, "count")}
                if "QTDE_CONSULTIVO" in df_cons_f.columns:
                    agrup_vend["Qtde Consultivos"] = ("QTDE_CONSULTIVO", "sum")
                if "QTDE_PRODUTOS" in df_cons_f.columns:
                    agrup_vend["Qtde Produtos"] = ("QTDE_PRODUTOS", "sum")
                if "QTDE_TV" in df_cons_f.columns:
                    agrup_vend["TV"] = ("QTDE_TV", "sum")
                if "QTDE_VIRTUA" in df_cons_f.columns:
                    agrup_vend["Virtua"] = ("QTDE_VIRTUA", "sum")
                if "QTDE_MESH" in df_cons_f.columns:
                    agrup_vend["Mesh/Outros"] = ("QTDE_MESH", "sum")
                if col_cons_valor:
                    agrup_vend["Receita"] = (col_cons_valor, "sum")

                df_vend_rank = (
                    df_cons_f.groupby(group_keys, dropna=False)
                    .agg(**agrup_vend)
                    .reset_index()
                )
                sort_col = (
                    "Qtde Consultivos"
                    if "Qtde Consultivos" in df_vend_rank.columns
                    else "Registros"
                )
                df_vend_rank = df_vend_rank.sort_values(
                    sort_col, ascending=False
                ).reset_index(drop=True)
                df_vend_rank.insert(0, "Posição", range(1, len(df_vend_rank) + 1))
                df_vend_rank["Posição"] = df_vend_rank["Posição"].apply(
                    Utilitarios.formatar_posicao
                )

                col_config_vend = {
                    "Posição": st.column_config.TextColumn("Pos."),
                    col_cons_tec: st.column_config.TextColumn("🔑 Login"),
                    "Registros": st.column_config.NumberColumn(
                        "📝 Registros", format="%d"
                    ),
                }
                if col_cons_nome and col_cons_nome in df_vend_rank.columns:
                    col_config_vend[col_cons_nome] = st.column_config.TextColumn(
                        "👤 Nome do Técnico"
                    )
                if "Qtde Consultivos" in df_vend_rank.columns:
                    col_config_vend["Qtde Consultivos"] = st.column_config.NumberColumn(
                        "💼 Consultivos", format="%d"
                    )
                if "Qtde Produtos" in df_vend_rank.columns:
                    col_config_vend["Qtde Produtos"] = st.column_config.NumberColumn(
                        "📦 Produtos", format="%d"
                    )
                if "TV" in df_vend_rank.columns:
                    col_config_vend["TV"] = st.column_config.NumberColumn(
                        "📺 TV", format="%d"
                    )
                if "Virtua" in df_vend_rank.columns:
                    col_config_vend["Virtua"] = st.column_config.NumberColumn(
                        "🌐 Virtua", format="%d"
                    )
                if "Mesh/Outros" in df_vend_rank.columns:
                    col_config_vend["Mesh/Outros"] = st.column_config.NumberColumn(
                        "📡 Mesh", format="%d"
                    )
                if "Receita" in df_vend_rank.columns:
                    col_config_vend["Receita"] = st.column_config.NumberColumn(
                        "💰 Receita (R$)", format="R$ %.2f"
                    )

                st.dataframe(
                    df_vend_rank,
                    use_container_width=True,
                    hide_index=True,
                    column_config=col_config_vend,
                    height=420,
                )

                top10_vend = df_vend_rank.head(10).copy()
                if col_cons_nome and col_cons_nome in top10_vend.columns:
                    top10_vend["Label"] = (
                        top10_vend[col_cons_tec].astype(str)
                        + " · "
                        + top10_vend[col_cons_nome].astype(str)
                    )
                else:
                    top10_vend["Label"] = top10_vend[col_cons_tec].astype(str)

                fig_vend = go.Figure(
                    go.Bar(
                        x=top10_vend[sort_col],
                        y=top10_vend["Label"],
                        orientation="h",
                        marker_color="#012869",
                        text=top10_vend[sort_col],
                        textposition="outside",
                    )
                )
                fig_vend.update_layout(
                    title=f"🥇 Top 10 Vendedores por {sort_col}",
                    margin=dict(l=0, r=60, t=40, b=0),
                    yaxis=dict(autorange="reversed"),
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    height=430,
                )
                st.plotly_chart(
                    fig_vend, use_container_width=True, config={"displayModeBar": False}
                )
                st.divider()

            if col_cons_data:
                st.subheader("📅 Evolução Temporal dos Consultivos")
                agrup_data = {"Registros": (col_cons_data, "count")}
                if "QTDE_CONSULTIVO" in df_cons_f.columns:
                    agrup_data["Consultivos"] = ("QTDE_CONSULTIVO", "sum")
                if "QTDE_PRODUTOS" in df_cons_f.columns:
                    agrup_data["Produtos"] = ("QTDE_PRODUTOS", "sum")

                df_temporal = (
                    df_cons_f.groupby(col_cons_data)
                    .agg(**agrup_data)
                    .reset_index()
                    .sort_values(col_cons_data)
                )

                fig_evol_cons = go.Figure()
                if "Consultivos" in df_temporal.columns:
                    fig_evol_cons.add_trace(
                        go.Scatter(
                            x=df_temporal[col_cons_data],
                            y=df_temporal["Consultivos"],
                            mode="lines+markers",
                            name="Consultivos",
                            line=dict(color="#012869", width=3),
                            marker=dict(size=8),
                            fill="tozeroy",
                            fillcolor="rgba(1,40,105,0.1)",
                        )
                    )
                if "Produtos" in df_temporal.columns:
                    fig_evol_cons.add_trace(
                        go.Scatter(
                            x=df_temporal[col_cons_data],
                            y=df_temporal["Produtos"],
                            mode="lines+markers",
                            name="Produtos",
                            line=dict(color="#F97316", width=3, dash="dot"),
                            marker=dict(size=8),
                        )
                    )
                fig_evol_cons.update_layout(
                    title="Volume Diário",
                    xaxis_title="Data",
                    yaxis_title="Quantidade",
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    margin=dict(l=0, r=20, t=40, b=0),
                    height=400,
                    hovermode="x unified",
                )
                st.plotly_chart(
                    fig_evol_cons,
                    use_container_width=True,
                    config={"displayModeBar": False},
                )

            st.divider()
            st.markdown("##### 📥 Exportar Dados Consultivos")
            abas_export = {"Consultivos_Filtrados": df_cons_f}
            if col_cons_tec and not df_vend_rank.empty:
                abas_export["Ranking_Vendedores"] = df_vend_rank
            excel_cons = Utilitarios.exportar_excel(abas_export)
            st.download_button(
                label="📊 Baixar Base Consultiva (XLSX)",
                data=excel_cons,
                file_name=f"consultivo_totale_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )

# ====================================================
# ── TAB 6: CRUD DE LOGINS ──
# ====================================================
with tab_logins:
    render_section_header("🔐", "Gerenciamento de Logins")
    try:
        db = get_db()
        df_users = db.get_users_dataframe()
        df_users = df_users.reindex(columns=HEADERS_USERS, fill_value="")
    except Exception as exc:
        st.error(f"Não foi possível carregar a base de logins: {exc}")
        df_users = pd.DataFrame(columns=HEADERS_USERS)

    colunas_publicas = ["Técnico", "Login", "User", "Perfil"]
    st.dataframe(df_users[colunas_publicas], use_container_width=True, hide_index=True)

    st.subheader("Adicionar login")
    with st.form("form_novo_login", clear_on_submit=True):
        novo_tecnico, novo_login, novo_user = st.columns(3)
        with novo_tecnico:
            tecnico = st.text_input("Técnico")
        with novo_login:
            login = st.text_input("Login")
        with novo_user:
            user = st.text_input("User")
        nova_senha, novo_perfil = st.columns(2)
        with nova_senha:
            senha = st.text_input("Senha", type="password")
        with novo_perfil:
            perfil = st.selectbox("Perfil", ["TÉCNICO", "ADMIN"])
        criar = st.form_submit_button("Criar login", type="primary")

    if criar:
        login_limpo = login.strip()
        user_limpo = user.strip()
        if not tecnico.strip() or not login_limpo or not user_limpo or not senha:
            st.error("Preencha Técnico, Login, User e Senha.")
        elif len(senha.strip()) < 6:
            st.error("A senha deve possuir no mínimo 6 caracteres.")
        elif (
            df_users["Login"]
            .astype(str)
            .str.strip()
            .str.upper()
            .eq(login_limpo.upper())
            .any()
            or df_users["User"]
            .astype(str)
            .str.strip()
            .str.upper()
            .eq(user_limpo.upper())
            .any()
        ):
            st.error("Login ou User já cadastrado.")
        else:
            novo_registro = {
                "Técnico": tecnico.strip(),
                "Login": login_limpo,
                "User": user_limpo,
                "Pass": bcrypt.hashpw(
                    senha.strip().encode("utf-8"), bcrypt.gensalt(rounds=12)
                ).decode("utf-8"),
                "Perfil": perfil,
            }
            df_gravar = pd.concat(
                [df_users, pd.DataFrame([novo_registro])], ignore_index=True
            )
            if db.save_users_dataframe(df_gravar):
                st.success("Login criado com sucesso.")
                st.rerun()
            else:
                st.error("Não foi possível salvar o novo login.")

    if not df_users.empty:
        st.subheader("Editar ou excluir login")
        logins = df_users["Login"].astype(str).tolist()
        login_selecionado = st.selectbox("Login cadastrado", logins)
        indice = df_users.index[df_users["Login"].astype(str) == login_selecionado][0]
        usuario_atual = df_users.loc[indice]

        with st.form("form_editar_login"):
            edit_tecnico, edit_user, edit_perfil = st.columns(3)
            with edit_tecnico:
                tecnico_editado = st.text_input(
                    "Técnico", value=str(usuario_atual["Técnico"])
                )
            with edit_user:
                user_editado = st.text_input("User", value=str(usuario_atual["User"]))
            with edit_perfil:
                perfil_atual = str(usuario_atual["Perfil"]).strip().upper()
                opcoes_perfil = ["TÉCNICO", "SUPERVISOR", "ADMIN"]
                perfil_editado = st.selectbox(
                    "Perfil",
                    opcoes_perfil,
                    index=(
                        opcoes_perfil.index(perfil_atual)
                        if perfil_atual in opcoes_perfil
                        else 0
                    ),
                )
            senha_editada = st.text_input(
                "Nova senha (deixe vazio para manter)", type="password"
            )
            salvar_edicao = st.form_submit_button("Salvar alterações")

        excluir = st.button("Excluir login selecionado", type="secondary")

        if salvar_edicao:
            user_editado_limpo = user_editado.strip()
            outros_users = df_users.index[
                (
                    df_users["User"].astype(str).str.strip().str.upper()
                    == user_editado_limpo.upper()
                )
                & (df_users.index != indice)
            ]
            if not tecnico_editado.strip() or not user_editado_limpo:
                st.error("Técnico e User são obrigatórios.")
            elif len(outros_users) > 0:
                st.error("User já cadastrado para outro login.")
            elif senha_editada and len(senha_editada.strip()) < 6:
                st.error("A nova senha deve possuir no mínimo 6 caracteres.")
            else:
                df_users.loc[indice, "Técnico"] = tecnico_editado.strip()
                df_users.loc[indice, "User"] = user_editado_limpo
                df_users.loc[indice, "Perfil"] = perfil_editado
                if senha_editada:
                    df_users.loc[indice, "Pass"] = bcrypt.hashpw(
                        senha_editada.strip().encode("utf-8"),
                        bcrypt.gensalt(rounds=12),
                    ).decode("utf-8")
                if db.save_users_dataframe(df_users):
                    st.success("Login atualizado com sucesso.")
                    st.rerun()
                else:
                    st.error("Não foi possível salvar as alterações.")

        if excluir:
            if login_selecionado.strip().upper() == login_logado.strip().upper():
                st.error("Não é permitido excluir o login atualmente utilizado.")
            else:
                df_restante = df_users.drop(index=indice).reset_index(drop=True)
                if db.save_users_dataframe(df_restante):
                    st.success("Login excluído com sucesso.")
                    st.rerun()
                else:
                    st.error("Não foi possível excluir o login.")